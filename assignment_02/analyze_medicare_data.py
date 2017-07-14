
# coding: utf-8

# In[1]:


# Importing the necessary library #

import requests
import os
import zipfile
import openpyxl
import sqlite3
import glob
import string
import csv
import sys
import xlsxwriter


# In[2]:

# ************************** CSV FILE PART ***************************************** #

# Link To Download the CSV Flat Files #

CSV_URL="https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"

# Request the response of the CSV_URL #

r=requests.get(CSV_URL)

# Create a 'Staging' Directory #

staging_dir_name="staging"
os.mkdir(staging_dir_name)

# Create a Path for the ZipFile #

zip_file_name=os.path.join(staging_dir_name,"DownLoadCSV.zip")

# Create "DownLoadCSV.zip" and the write the contents of the URL in the ZipFile #

ZipFile=open(zip_file_name,"wb")
ZipFile.write(r.content)
ZipFile.close()

# To UnZip all the CSV Files in the Staging table #

z=zipfile.ZipFile(zip_file_name,'r')
z.extractall(staging_dir_name)
z.close()


# In[3]:

# ************************** EXCEL FILE PART ***************************************** #

# Link To Download the EXCEL File #

EXCEL_URL="http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"

# Request the response of the EXCEL_URL #

r=requests.get(EXCEL_URL)

# Create "hospital_ranking_focus_states.xlsx" and the write the contents of the URL in the ZipFile #

xf=open("hospital_ranking_focus_states.xlsx","wb")
xf.write(r.content)
xf.close()


# In[4]:

## Open hospital_ranking_focus_states.xlsx ##

wb=openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")

## Print the Sheets Names in the WorkBook  ##


for sheet in wb.get_sheet_names():
    print(sheet)
    
       
## Assign Sheet Names to respective sheets ##

sheet1=wb.get_sheet_by_name("Hospital National Ranking")
sheet2=wb.get_sheet_by_name("Focus States")



i=2
StatesList=[]
while sheet2.cell(row=i,column=1).value!=None:
    #print(sheet2.cell(row=i,column=1).value,"|",sheet2.cell(row=i,column=2).value)
    State=[sheet2.cell(row=i,column=1).value]
    StatesList=StatesList+State
    i+=1

StatesList.insert(0, "Nationwide")


# In[5]:

#wb=openpyxl.Workbook()

#wb.remove_sheet(wb.get_sheet_by_name('Sheet'))



#Sheets=0

#while Sheets<11:
    
#    sheet_1=wb.create_sheet(StatesList[Sheets])
#    i=1
#    head=0
#    Headers=["Provider ID","Hospital Name","City","State","County"]
#    while i<6:

#        sheet_1.cell(row=1,column=i,value=Headers[head])
#        i=i+1
#        head=head+1
#    Sheets+=1
    
    ## Save the file as hospital_ranking.xlsx ##

#wb.save("hospital_ranking.xlsx")

#wb=openpyxl.Workbook()

#wb.remove_sheet(wb.get_sheet_by_name('Sheet'))



#Sheets=0

#while Sheets<11:
    
#    sheet_1=wb.create_sheet(StatesList[Sheets])
#    i=1
#    head=0
#    Headers=["Provider ID","Hospital Name","City","State","County"]
#    while i<6:

#        sheet_1.cell(row=1,column=i,value=Headers[head])
#        i=i+1
#        head=head+1
#    Sheets+=1
    
    ## Save the file as hospital_ranking.xlsx ##

#wb.save("hospital_ranking.xlsx")


#wb=openpyxl.Workbook()

#wb.remove_sheet(wb.get_sheet_by_name('Sheet'))



#Sheets=0

#while Sheets<11:
    
#    sheet_1=wb.create_sheet(StatesList[Sheets])
#    i=1
#    head=0
#    Headers=["Provider ID","Hospital Name","City","State","County"]
#    while i<6:

#        sheet_1.cell(row=1,column=i,value=Headers[head])
#        i=i+1
#        head=head+1
#    Sheets+=1
    
    ## Save the file as hospital_ranking.xlsx ##

#wb.save("hospital_ranking.xlsx")





# In[6]:

#wb=openpyxl.Workbook()

#wb.remove_sheet(wb.get_sheet_by_name('Sheet'))



#Sheets=0

#while Sheets<11:
    
#    sheet_1=wb.create_sheet(StatesList[Sheets])
#    i=1
#    head=0
#    Headers=["Measure ID","Measure Name","Minimum","Maximum","Average","Standard Deviation"]
#    while i<7:

#        sheet_1.cell(row=1,column=i,value=Headers[head])
#        i=i+1
#        head=head+1
#    Sheets+=1
    
    ## Save the file as measures_statistics.xlsx ##

#wb.save("measures_statistics.xlsx")




# In[7]:

## Create an SQL Connection as medicare_hospital_compare.db ##

conn=sqlite3.connect("medicare_hospital_compare.db")


# In[8]:

## Getting the list of files and creating a list based on ext .CSV,.XLSX ##

glob_dir=os.path.join(staging_dir_name,"*.csv")
CSVFileList_With_CSV_EXT=[]
CSVFileList_Without_EXT=[]
CSVFileList_With_XLSX_EXT=[]
for file_name in glob.glob(glob_dir):
    #print(file_name)
    CSVFileList_With_CSV_EXT=CSVFileList_With_CSV_EXT+[(os.path.basename(file_name))]
    CSVFileList_Without_EXT=CSVFileList_Without_EXT+[(os.path.splitext(os.path.basename(file_name))[0])]
    CSVFileList_With_XLSX_EXT=CSVFileList_With_XLSX_EXT+[(os.path.splitext(os.path.basename(file_name))[0])+".xlsx"]
    #print("       directory name",os.path.dirname(file_name))
    #print("          split ext",os.path.splitext(os.path.basename(file_name)))
    #print("              absolute path:",os.path.abspath(file_name))


# In[9]:

## Commands to remove the Corrupted file - FY2015_Percent_Change_in_Medicare_Payments from the file list ##

CSVFileList_Without_EXT.remove("FY2015_Percent_Change_in_Medicare_Payments")
CSVFileList_With_CSV_EXT.remove("FY2015_Percent_Change_in_Medicare_Payments.csv")
CSVFileList_With_XLSX_EXT.remove("FY2015_Percent_Change_in_Medicare_Payments.xlsx")


# In[14]:

## Commands to do the encoding part ##

fn=os.path.join(staging_dir_name,"FY2015_Percent_Change_in_Medicare_Payments.csv")
in_fp=open(fn,"rt",encoding='cp1252')
input_data=in_fp.read()
in_fp.close()


# In[15]:

## commands to do the decoding part ##

ofn=os.path.join(staging_dir_name,"FY2015_Percent_Change_in_Medicare_Payments.csv")
out_fp=open(ofn,"wt",encoding='utf-8')
for c in input_data:
    if c!='\0':
        out_fp.write(c)
out_fp.close()


# In[16]:


## Code snippet to convert the .CSV files to .Xlsx files ##

if __name__ == '__main__':
    #listOfFiles = os.listdir(directory)           #  list of all files in the directory
    listOfFiles = glob.glob(glob_dir)                       
    for index, fileInList in enumerate(listOfFiles):     
        fileName  = fileInList[0:fileInList.find('.csv')]     
        excelFile = xlsxwriter.Workbook(fileName + '.xlsx')
        worksheet = excelFile.add_worksheet()    
        #with open(fileName + ".csv", 'rb') as f:
        #print("File Name --> ",fileName)
        with open(fileInList, 'rt',encoding='cp1252') as f:   
            content = csv.reader(f)
            for index_row, data_in_row in enumerate(content):
                for index_col, data_in_cell in enumerate(data_in_row):
                    worksheet.write(index_row, index_col, data_in_cell)
                    #print("File Name Created --> ",fileName)
        #print("File Name Created --> ",fileName)
    excelFile.close()
   # print (" === Conversion is done ===")


# In[17]:

## Function to return the table names after the condition/rules are applied ##

def TableName_Transformation(OldValue):
    NewValue=OldValue.lower()
    NewValue=NewValue.replace(" ", "_")
    NewValue=NewValue.replace("-", "_")
    NewValue=NewValue.replace("%", "pct")
    NewValue=NewValue.replace("/", "_")
    i=0
    while i<len(list(string.ascii_lowercase)):
        if NewValue[0] == list(string.ascii_lowercase)[i]:
            return NewValue;
        i+=1
    else:
        return ("t_"+NewValue);


# In[18]:

## Function to return the column names after the condition/rules are applied ##

def ColumnName_Transformation(OldValue):
    NewValue=OldValue.lower()
    NewValue=NewValue.replace(" ", "_")
    NewValue=NewValue.replace("-", "_")
    NewValue=NewValue.replace("%", "pct")
    NewValue=NewValue.replace("/", "_")
    i=0
    while i<len(list(string.ascii_lowercase)):
        if NewValue[0] == list(string.ascii_lowercase)[i]:
            return NewValue;
        i+=1
    else:
        return ("t_"+NewValue);


# In[20]:

conn=sqlite3.connect("medicare_hospital_compare.db")
c1=conn.cursor()
c2=conn.cursor()
FileNo=0
while FileNo<61:
    wb=openpyxl.load_workbook(os.path.join(staging_dir_name,CSVFileList_With_XLSX_EXT[FileNo]))
    sheet=wb.get_sheet_by_name("Sheet1")
    TableColumns=[]
    i=1
    while sheet.cell(row=1,column=i).value!=None:
        Columns=[sheet.cell(row=1,column=i).value]
        TableColumns+=Columns
        i+=1
    CreateStr = "CREATE TABLE IF NOT EXISTS " + TableName_Transformation(CSVFileList_Without_EXT[FileNo]) + "(" 
    j=0
    CombinedColumn=""
    CombinedColumn_ins=[]
    while j<(len(TableColumns)):
        if j<(len(TableColumns)-1):
            SingleColumn=(ColumnName_Transformation(TableColumns[j])+ " TEXT, ")
            SingleColumn_ins=[(ColumnName_Transformation(TableColumns[j]))]
        else:
            SingleColumn=(ColumnName_Transformation(TableColumns[j])+ " TEXT")
            SingleColumn_ins=[ColumnName_Transformation(TableColumns[j])]
        CombinedColumn=CombinedColumn+SingleColumn
        CombinedColumn_ins=CombinedColumn_ins+SingleColumn_ins
        j+=1
    
    SQL_QUERY=CreateStr+CombinedColumn+")"
    #print(SQL_QUERY)
    c1.execute(SQL_QUERY)

    
    
    tup = []
    for y, rows in enumerate(sheet):
        tuprow = []
        if y == 0:
            continue
        for row in rows:
            if str(row.value).strip() != 'None':
                tuprow.append(str(row.value).strip()) 
                #tuprow=tuprow+(str(row.value).strip())
            else:
                tuprow.append('')
                #tupnew=
        tuprow.pop(-1)
        tup.append(tuprow)
        #tup=tup+tuple(tuprow)
        #print("loop end")
    #print("full loop ends")
    tuple_1=[tuple(l) for l in tup]
    insQuery1 = 'INSERT INTO '+ TableName_Transformation(CSVFileList_Without_EXT[FileNo]) + "("
    insQuery2 = ''
    for col in CombinedColumn_ins:
        insQuery1 += col + ', '
        insQuery2 += '?, '
    insQuery1 = insQuery1[:-2] + ') VALUES('
    insQuery2 = insQuery2[:-2] + ')'
    #print("start")
    insQuery = insQuery1 + insQuery2
    #print("end")
    c2.executemany(insQuery,tuple(tuple_1))
    conn.commit()
    FileNo+=1


# In[21]:

conn=sqlite3.connect("medicare_hospital_compare.db")
c1=conn.cursor()
c2=conn.cursor()
FileNo=0
while FileNo<1:
    wb=openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")
    sheet=wb.get_sheet_by_name("Hospital National Ranking")
    TableColumns=[]
    i=1
    while sheet.cell(row=1,column=i).value!=None:
        Columns=[sheet.cell(row=1,column=i).value]
        TableColumns+=Columns
        i+=1
    CreateStr = "CREATE TABLE IF NOT EXISTS " + TableName_Transformation("Hospital National Ranking") + "(" 
    j=0
    CombinedColumn=""
    CombinedColumn_ins=[]
    while j<(len(TableColumns)):
        if j<(len(TableColumns)-1):
            SingleColumn=(ColumnName_Transformation(TableColumns[j])+ " TEXT, ")
            SingleColumn_ins=[(ColumnName_Transformation(TableColumns[j]))]
        else:
            SingleColumn=(ColumnName_Transformation(TableColumns[j])+ " NUMBER")
            SingleColumn_ins=[ColumnName_Transformation(TableColumns[j])]
        CombinedColumn=CombinedColumn+SingleColumn
        CombinedColumn_ins=CombinedColumn_ins+SingleColumn_ins
        j+=1
    
    SQL_QUERY=CreateStr+CombinedColumn+")"
    #print(SQL_QUERY)
    c1.execute(SQL_QUERY)

    
    
    tup = []
    for y, rows in enumerate(sheet):
        tuprow = []
        if y == 0:
            continue
        for row in rows:
            if str(row.value).strip() != 'None':
                tuprow.append(str(row.value).strip()) 
                #tuprow=tuprow+(str(row.value).strip())
            else:
                tuprow.append('')
                #tupnew=
        tuprow.pop(-1)
        tup.append(tuprow)
        #tup=tup+tuple(tuprow)
        #print("loop end")
    #print("full loop ends")
    tuple_1=[tuple(l) for l in tup]
    insQuery1 = 'INSERT INTO '+ TableName_Transformation("Hospital National Ranking") + "("
    insQuery2 = ''
    for col in CombinedColumn_ins:
        insQuery1 += col + ', '
        insQuery2 += '?, '
    insQuery1 = insQuery1[:-2] + ') VALUES('
    insQuery2 = insQuery2[:-2] + ')'
    #print("start")
    insQuery = insQuery1 + insQuery2
    #print("end")
    c2.executemany(insQuery,tuple(tuple_1))
    conn.commit()
    FileNo+=1


# In[22]:

StatesDictionary={'California':'CA','Florida':'FL','Georgia':'GA','Illinois':'IL','Kansas':'KS','Michigan':'MI','New York':'NY','Ohio':'OH','Pennsylvania':'PA','Texas':'TX'}


# In[24]:

import sqlite3
from xlsxwriter.workbook import Workbook
workbook = Workbook('hospital_ranking.xlsx')
SheetNum=0
c=conn.cursor()
while SheetNum<11:
    worksheet = workbook.add_worksheet(StatesList[SheetNum])
    
    if StatesList[SheetNum]=='Nationwide':
        mysel=c.execute("""select hospital_national_ranking.provider_id,hospital_name,city,state,county_name
        from hospital_general_information
        join hospital_national_ranking
        on hospital_national_ranking.provider_id=hospital_general_information.provider_id
        order by ranking limit 100""")
    else:
        l=[StatesDictionary[StatesList[SheetNum]]]
        mysel=c.execute("""select hospital_national_ranking.provider_id,hospital_name,city,state,county_name
        from hospital_general_information
        join hospital_national_ranking
        on hospital_national_ranking.provider_id=hospital_general_information.provider_id
        where state in ('""" + ','.join(map(str, l)) +"') order by ranking limit 100")
    Headers=["Provider ID","Hospital Name","City","State","County"]
    j=0
    while j<5:
        worksheet.write(0, j,Headers[j])
        j+=1
    i=10
    for i, row in enumerate(mysel):
        #print (row)
        worksheet.write(i+1, 0, row[0])
        worksheet.write(i+1, 1, row[1])
        worksheet.write(i+1, 2, row[2])
        worksheet.write(i+1, 3, row[3])
        worksheet.write(i+1, 4, row[4])
    SheetNum+=1
workbook.close()


# In[26]:

import sqlite3
from xlsxwriter.workbook import Workbook
workbook = Workbook('measures_statistics.xlsx')
SheetNum=0
c=conn.cursor()
while SheetNum<11:
    worksheet = workbook.add_worksheet(StatesList[SheetNum])
    
 #   if StatesList[SheetNum]=='Nationwide':
 #       mysel=c.execute("""select hospital_national_ranking.provider_id,hospital_name,city,state,county_name
 #       from hospital_general_information
 #       join hospital_national_ranking
 #       on hospital_national_ranking.provider_id=hospital_general_information.provider_id
 #       order by ranking limit 100""")
 #   else:
 #       l=[StatesDictionary[StatesList[SheetNum]]]
 #       mysel=c.execute("""select hospital_national_ranking.provider_id,hospital_name,city,state,county_name
 #       from hospital_general_information
 #       join hospital_national_ranking
 #       on hospital_national_ranking.provider_id=hospital_general_information.provider_id
 #       where state in ('""" + ','.join(map(str, l)) +"') order by ranking limit 100")
    Headers=["Measure ID","Measure Name","Minimum","Maximum","Average","Standard Deviation"]
    j=0
    while j<6:
        worksheet.write(0, j,Headers[j])
        j+=1
 #   i=10
 #   for i, row in enumerate(mysel):
        #print (row)
 #       worksheet.write(i+1, 0, row[0])
 #       worksheet.write(i+1, 1, row[1])
 #       worksheet.write(i+1, 2, row[2])
 #       worksheet.write(i+1, 3, row[3])
 #       worksheet.write(i+1, 4, row[4])
    SheetNum+=1
workbook.close()


# In[ ]:



